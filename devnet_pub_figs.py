#!/usr/bin/env python3
# SPDX-License-Identifier: Apache-2.0
#
# Copyright 2026 ZeroNode
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
# http://www.apache.org/licenses/LICENSE-2.0
#
# devnet_pub_figs.py
#
# Objective
# - Generate publication-ready Figures and Tables for the Datacenter BYOG
#   research workflow.
#
# Architecture
# - devnetDC_sysdsg.xlsx is the experiment and publication-definition driver.
# - devnet_pub_figs.py is a generic parser and renderer.
# - Scenario-specific stress_out/index.html files provide committed OPF results.
#
# What it does
# - Reads active Scenario definitions from devnetDC_sysdsg.xlsx.
# - Ignores Scenarios prefixed with [NA].
# - Reads XLSX-defined Event groupings, Commit IDs and Event Conditions.
# - Reads XLSX-defined publication Artifact type, number and filename suffix.
# - Resolves the Scenario-specific stress_out/index.html result source.
# - Maps each Event Commit ID to its corresponding solved OPF result.
# - Generates XLSX-defined publication Figures and Tables.
# - Renders 8760-hour chronology Figures showing:
#     - Objective / system feasibility
#     - Total system load
#     - Datacenter p_set and BYOG p_nom
#     - Percentage-valued Event Driver conditions
# - Marks infeasible intervals with a dark-red bar.
#
# Event Driver Visual Convention
# - Adverse grid/system drivers are plotted Dark Red when the Event Condition
#   contains: failure, outage, constraint, deration, or congestion.
# - Datacenter mitigation/support drivers are plotted Blue.
# - Event Condition wording is therefore part of the plotting interface and
#   should remain consistent in devnetDC_sysdsg.xlsx.
#
# Inputs
# - ./devnet-stress-vectors/devnetDC_sysdsg.xlsx
#     - Worksheet: devnetDC_dsg
#     - Scenario / Artifact / Event definitions
#     - Scenario-specific index.html paths
# - Scenario-specific stress_out/index.html files.
#
# Outputs
# - ./pub_figs/
# - Output filename, Artifact type and Artifact number are XLSX-defined.
#
# Run
#     python devnet_pub_figs.py
#
# Workflow
#     devnetDC_sysdsg.xlsx
#              |
#              +--> active Scenario
#              +--> Artifact definition
#              +--> Event grouping
#              +--> Commit ID
#              +--> index.html result
#                         |
#                         v
#                 devnet_pub_figs.py
#                         |
#                         v
#                    ./pub_figs/
# ------------------------------------------------------------------------------

# ------------------------------------------------------------------------------
# --- Standard library imports ---
# ------------------------------------------------------------------------------
import argparse
import html
import math
import re
import sys
import textwrap
from pathlib import Path

# ------------------------------------------------------------------------------
# --- Third-party imports ---
# ------------------------------------------------------------------------------
import matplotlib.pyplot as plt
import openpyxl

# ------------------------------------------------------------------------------
# --- Configuration: paths, filenames, worksheet selection ---
# ------------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX = SCRIPT_DIR / "devnet-stress-vectors" / "devnetDC_sysdsg.xlsx"
DEFAULT_OUTDIR = SCRIPT_DIR / "pub_figs"

EVENT_SHEET_NAME = "devnetDC_dsg"

NA_PREFIX = "[NA]"         # Not Applicable Scenario prefix in devnetDC_sysdsg.xlsx

REPORT_TAG_COLUMN = 3      # C :: devnet_stress.py > o/p:
REPORT_PATH_COLUMN = 5     # E :: relative index.html path

DPI = 300
HOURS_PER_YEAR = 8760

# ------------------------------------------------------------------------------
# devnetDC_dsg worksheet column mapping.
#
# These constants define the interface between devnetDC_sysdsg.xlsx and this
# renderer. If the worksheet layout changes, update the mappings here rather
# than introducing column numbers inside parsing functions.
# ------------------------------------------------------------------------------
EVENT_TAG_COLUMN = 4       # D :: Event
COMMIT_TAG_COLUMN = 5      # E :: Commit ID
CONDITION_COLUMN = 6       # F :: Event Conditions
CONDITION_VALUE_COLUMN = 7 # G :: Failure / mitigation value
START_HOUR_COLUMN = 8      # H :: Start hour
END_HOUR_COLUMN = 9        # I :: End hour

SCENARIO_TAG_COLUMN = 3    # C :: "Scenario"
SCENARIO_TEXT_COLUMN = 4   # D :: Scenario heading

ARTIFACT_EVENTS_COLUMN = 3 # C :: Event grouping
ARTIFACT_TYPE_COLUMN = 4   # D :: Fig / Table
ARTIFACT_NUMBER_COLUMN = 5 # E :: Publication number
ARTIFACT_SUFFIX_COLUMN = 6 # F :: Output filename suffix

ASSET_TAG_COLUMN = 16      # P
ASSET_P_NOM_COLUMN = 18    # R
ASSET_MC_COLUMN = 19       # S
ASSET_LOAD_COLUMN = 20     # T

# ------------------------------------------------------------------------------
# DevNet buses represented in publication operating-state tables.
#
# Bus asset parameters are read from the XLSX Event panel. PJM_NE and SERC_SE
# are currently highlighted in publication tables as the critical neighboring
# buses investigated by the reliability/congestion experiments.
# ------------------------------------------------------------------------------
BUSES = [
    "WECC_NW",
    "WECC_SW",
    "SPP_MISO",
    "PJM_NE",
    "SERC_SE",
    "ERCOT",
]

# ------------------------------------------------------------------------------
# --- Command-line argument parsing ---
# ------------------------------------------------------------------------------
# ------------------------------------------------------------------------------
# parse_args()
# Parses command-line overrides for workbook, stress report and output directory.
#
# --xlsx   Overrides the default devnetDC_sysdsg.xlsx path.
# --html   Optionally overrides the Scenario-specific XLSX index.html path.
# --outdir Overrides the default ./pub_figs output directory.
# ------------------------------------------------------------------------------
def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    # Optional CLI override. If omitted, index.html is resolved from
    # the active Scenario definition in devnetDC_sysdsg.xlsx.
    p.add_argument("--html", default=None)
    p.add_argument("--outdir", default=str(DEFAULT_OUTDIR))
    return p.parse_args()

# ------------------------------------------------------------------------------
# resolve_report_path()
# Resolves the Scenario-specific stress_out/index.html path defined in XLSX.
#
# Supports:
# - Script-relative paths beginning with ~\ or ~/
# - Absolute paths
# - Paths relative to the workbook directory
# - XLSX notation using "\:" or "/:"
#
# Allows different Scenarios to source committed OPF results from different
# DevNet stress reports without hard-coded report paths in this script.
# ------------------------------------------------------------------------------
def resolve_report_path(
    report_path,
    workbook_path,
):

    if report_path is None:
        raise ValueError(
            "No devnet_stress.py output path defined for Scenario."
        )

    raw = str(
        report_path
    ).strip()

    # ------------------------------------------------------------------
    # Support workbook notation:
    #   ~\devnetDC-sld\stress_out\:index.html
    #
    # "~\" means relative to the pypsa-zn / script directory.
    # "\:" is normalized to "\" before resolving the path.
    # ------------------------------------------------------------------
    raw = raw.replace(
        "\\:",
        "\\",
    )

    raw = raw.replace(
        "/:",
        "/",
    )

    if raw.startswith(
        ("~\\", "~/")
    ):

        relative = raw[2:]

        relative = relative.replace(
            "\\",
            "/",
        )

        return (
            SCRIPT_DIR
            / relative
        ).resolve()

    normalized = raw.replace(
        "\\",
        "/",
    )

    path = Path(
        normalized
    )

    if path.is_absolute():
        return path.resolve()

    return (
        workbook_path.parent
        / path
    ).resolve()

# ------------------------------------------------------------------------------
# confirm()
# Returns True when the user explicitly confirms an interactive Y/N prompt.
# ------------------------------------------------------------------------------
def confirm(prompt):
    ans = input(f"{prompt} (Y/N): ").strip().lower()
    return ans in ("y", "yes")

# ------------------------------------------------------------------------------
# prepare_output_dir()
# Validates and prepares the publication output directory.
#
# Requires user confirmation before:
# - Creating ./pub_figs when it does not exist.
# - Writing new publication artifacts into an existing ./pub_figs directory.
# ------------------------------------------------------------------------------
def prepare_output_dir(outdir: Path) -> Path:
    if outdir.exists():
        if not outdir.is_dir():
            raise ValueError(
                f"Publication output path exists but is not a directory:\n\t{outdir}"
            )

        if not confirm(
            f"pub_figs directory already exists:\n\t{outdir}\n"
            "Overwrite publication figures in this directory?"
        ):
            print("Publication figure generation cancelled.")
            sys.exit(0)

    else:
        if not confirm(
            f"pub_figs directory does not exist:\n\t{outdir}\n"
            "Create it?"
        ):
            print("Publication figure generation cancelled.")
            sys.exit(0)

        outdir.mkdir(parents=True, exist_ok=True)

    return outdir

# ------------------------------------------------------------------------------
# load_events()
# Parses active Event definitions from the devnetDC_dsg worksheet.
#
# For each Event:
# - Associates it with the nearest preceding Scenario.
# - Ignores Events belonging to an [NA] Scenario.
# - Reads Commit ID, Event Conditions and Start/End hour.
# - Reads six-bus asset p_nom, marginal cost and load parameters.
# - Reads DC_PJM_NE p_nom, marginal cost and p_set.
#
# Event 0 is stored internally as "baseline"; Event N is stored as "eventN".
# ------------------------------------------------------------------------------
def load_events(path):
    wb = openpyxl.load_workbook(
        path,
        data_only=True,
    )

    if EVENT_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Event worksheet '{EVENT_SHEET_NAME}' "
            f"not found in {path.name}."
        )

    # --------------------------------------------------------------
    # Publication events are read only from devnetDC_dsg.
    # Other workbook sheets, such as EventHours, are working caches
    # and are intentionally ignored.
    # --------------------------------------------------------------
    ws = wb[
        EVENT_SHEET_NAME
    ]

    # --------------------------------------------------------------------
    # Locate event panels using "Event N" tags in column EVENT_TAG_COLUMN.
    # --------------------------------------------------------------------
    event_headers = []

    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, EVENT_TAG_COLUMN).value

        if isinstance(value, str):
            match = re.fullmatch(
                r"Event\s+(\d+)",
                value.strip(),
                re.IGNORECASE,
            )

            if match:
                event_headers.append(
                    (
                        int(match.group(1)),
                        row,
                    )
                )

    if not event_headers:
        raise ValueError(
            "No Event N panels found in devnetDC_sysdsg.xlsx."
        )

    events = {}

    # ------------------------------------------------------------------
    # Map each event panel to the nearest preceding Scenario heading.
    # ------------------------------------------------------------------
    scenario_rows = []

    for row in range(1, ws.max_row + 1):

        scenario_tag = ws.cell(
            row,
            SCENARIO_TAG_COLUMN,
        ).value

        scenario_text = ws.cell(
            row,
            SCENARIO_TEXT_COLUMN,
        ).value

        if (
            isinstance(scenario_tag, str)
            and scenario_text is not None
        ):

            tag = scenario_tag.strip()

            if (
                tag.lower() == "scenario"
                or tag.lower() == "[na] scenario"
            ):

                scenario = str(
                    scenario_text
                ).strip()

                active = not tag.lower().startswith(
                    NA_PREFIX.lower()
                )

                scenario_rows.append(
                    (
                        row,
                        scenario,
                        active,
                    )
                )

    # ------------------------------------------------------------------
    # Parse each event panel independently.
    # ------------------------------------------------------------------
    for index, (event_num, start_row) in enumerate(event_headers):

        end_row = (
            event_headers[index + 1][1] - 1
            if index + 1 < len(event_headers)
            else ws.max_row
        )

        key = "baseline" if event_num == 0 else f"event{event_num}"

        # --------------------------------------------------------------
        # Associate this event with its nearest preceding Scenario row.
        # --------------------------------------------------------------
        scenario = None
        scenario_active = True

        for (
            scenario_row,
            scenario_text,
            active,
        ) in scenario_rows:

            if scenario_row < start_row:
                scenario = scenario_text
                scenario_active = active
            else:
                break

        # --------------------------------------------------------------
        # Ignore all Events belonging to an [NA] Scenario.
        # --------------------------------------------------------------
        if not scenario_active:
            continue

        if scenario is None:
            scenario = f"Event {event_num}"

        commit = None
        start_hour = None
        end_hour = None
        conditions = []

        # --------------------------------------------------------------
        # Commit ID + event conditions + event window.
        # --------------------------------------------------------------
        for row in range(start_row, end_row + 1):

            # --------------------------------------------------------------
            # Read Commit ID assigned to this event in devnetDC_sysdsg.xlsx.
            # This provides the link between the experiment definition in
            # the workbook and its solved result in stress_out/index.html.
            # --------------------------------------------------------------
            commit_value = ws.cell(row, COMMIT_TAG_COLUMN).value

            if (
                isinstance(commit_value, str)
                and re.fullmatch(r"C\d+", commit_value.strip(), re.IGNORECASE)
            ):
                commit = commit_value.strip().lower()

            condition = ws.cell(row, CONDITION_COLUMN).value
            condition_value = ws.cell(
                row,
                CONDITION_VALUE_COLUMN,
            ).value

            if condition is not None:
                conditions.append(
                    {
                        "name": str(condition).strip(),
                        "value": condition_value,
                    }
                )

            sh = ws.cell(row, START_HOUR_COLUMN).value
            eh = ws.cell(row, END_HOUR_COLUMN).value

            if sh is not None and eh is not None:
                try:
                    start_hour = int(sh)
                    end_hour = int(eh)
                except (TypeError, ValueError):
                    pass

        if commit is None:
            raise ValueError(
                f"Commit ID not found for Event {event_num}."
            )

        # --------------------------------------------------------------
        # Parse devnet_assets block.
        # --------------------------------------------------------------
        assets = {}

        for row in range(start_row, end_row + 1):

            bus = ws.cell(row, ASSET_TAG_COLUMN).value

            if bus in BUSES:
                assets[str(bus)] = {
                    "p_nom": float(
                        ws.cell(
                            row,
                            ASSET_P_NOM_COLUMN,
                        ).value
                    ),
                    "mc": float(
                        ws.cell(
                            row,
                            ASSET_MC_COLUMN,
                        ).value
                    ),
                    "load": float(
                        ws.cell(
                            row,
                            ASSET_LOAD_COLUMN,
                        ).value
                    ),
                }

        missing_buses = [
            bus
            for bus in BUSES
            if bus not in assets
        ]

        if missing_buses:
            raise ValueError(
                f"Missing asset rows in Event {event_num}: "
                + ", ".join(missing_buses)
            )

        # --------------------------------------------------------------
        # Parse datacenter row dynamically.
        # --------------------------------------------------------------
        dc_row = None

        for row in range(start_row, end_row + 1):
            if (
                str(ws.cell(row, ASSET_TAG_COLUMN).value).strip()
                == "DC_PJM_NE"
            ):
                dc_row = row
                break

        if dc_row is None:
            raise ValueError(
                f"DC_PJM_NE row not found for Event {event_num}."
            )

        events[key] = {
            "event_num": event_num,
            "event_label": f"Event {event_num}",
            "scenario": scenario,
            "commit": commit,
            "start_hour": start_hour,
            "end_hour": end_hour,
            "conditions": conditions,
            "assets": assets,
            "dc_p_nom": float(
                ws.cell(
                    dc_row,
                    ASSET_P_NOM_COLUMN,
                ).value
            ),
            "dc_mc": float(
                ws.cell(
                    dc_row,
                    ASSET_MC_COLUMN,
                ).value
            ),
            "dc_p_set": float(
                ws.cell(
                    dc_row,
                    ASSET_LOAD_COLUMN,
                ).value
            ),
        }

    return events

# ------------------------------------------------------------------------------
# load_artifacts()
# Parses publication Artifact definitions from active XLSX Scenarios.
#
# An Artifact is processed only when its Scenario provides:
# - Scenario title
# - Scenario-specific stress-report path
# - Event grouping
# - Artifact type: Fig or Table
# - Artifact number
# - Output filename suffix
#
# Incomplete Artifact definitions and [NA] Scenarios are excluded from the
# publication-artifact selection menu.
# ------------------------------------------------------------------------------
def load_artifacts(path):

    wb = openpyxl.load_workbook(
        path,
        data_only=True,
    )

    if EVENT_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Event worksheet '{EVENT_SHEET_NAME}' "
            f"not found in {path.name}."
        )

    ws = wb[
        EVENT_SHEET_NAME
    ]

    artifacts = []
    current_scenario = None
    current_events = []
    current_report_path = None
    scenario_active = False

    for row in range(
        1,
        ws.max_row + 1,
    ):

        # --------------------------------------------------------------
        # Track the active Scenario heading.
        # --------------------------------------------------------------
        scenario_tag = ws.cell(
            row,
            SCENARIO_TAG_COLUMN,
        ).value

        scenario_text = ws.cell(
            row,
            SCENARIO_TEXT_COLUMN,
        ).value

        if (
            isinstance(scenario_tag, str)
            and scenario_text is not None
        ):

            tag = scenario_tag.strip()

            if (
                tag.lower() == "scenario"
                or tag.lower() == "[na] scenario"
            ):

                scenario_active = (
                    not tag.lower().startswith(
                        NA_PREFIX.lower()
                    )
                )

                current_scenario = (
                    str(
                        scenario_text
                    ).strip()
                    if scenario_active
                    else None
                )

                current_events = []
                current_report_path = None

                continue

        # --------------------------------------------------------------
        # Read devnet_stress.py report path associated with Scenario.
        # --------------------------------------------------------------
        report_tag = ws.cell(
            row,
            REPORT_TAG_COLUMN,
        ).value

        report_path = ws.cell(
            row,
            REPORT_PATH_COLUMN,
        ).value

        if (
            scenario_active
            and isinstance(
                report_tag,
                str,
            )
            and report_tag.strip().lower().startswith(
                "devnet_stress.py"
            )
            and report_path is not None
        ):
            current_report_path = str(
                report_path
            ).strip()

            continue

        # --------------------------------------------------------------
        # Read explicit Event grouping for this Scenario artifact.
        # --------------------------------------------------------------
        event_group = ws.cell(
            row,
            ARTIFACT_EVENTS_COLUMN,
        ).value

        artifact_type = ws.cell(
            row,
            ARTIFACT_TYPE_COLUMN,
        ).value

        artifact_number = ws.cell(
            row,
            ARTIFACT_NUMBER_COLUMN,
        ).value

        artifact_suffix = ws.cell(
            row,
            ARTIFACT_SUFFIX_COLUMN,
        ).value

        if event_group is not None:

            event_numbers = [
                int(value)
                for value in re.findall(
                    r"Event\s+(\d+)",
                    str(event_group),
                    re.IGNORECASE,
                )
            ]

            if event_numbers:
                current_events = event_numbers

        # --------------------------------------------------------------
        # Artifact rows are driven entirely by XLSX metadata.
        # Skip incomplete placeholder rows.
        # --------------------------------------------------------------
        if (
            not scenario_active
            or artifact_type is None
            or artifact_number is None
            or artifact_suffix is None
            or current_scenario is None
            or current_report_path is None
            or not current_events
        ):
            continue

        artifact_type = str(
            artifact_type
        ).strip().lower()

        if artifact_type not in (
            "fig",
            "table",
        ):
            continue

        artifacts.append(
            {
                "type": artifact_type,
                "number": int(
                    artifact_number
                ),
                "scenario": current_scenario,
                "report_path": current_report_path,
                "events": list(
                    current_events
                ),
                "suffix": str(
                    artifact_suffix
                ).strip(),
            }
        )

    return artifacts

# ------------------------------------------------------------------------------
# load_results()
# Parses committed OPF results from a DevNet stress_out/index.html report.
#
# Results are indexed by Commit ID (c1, c2, ...).
#
# Extracts:
# - Objective value
# - Total system load
# - Feasibility state
# - Per-bus net import/export
#
# Feasibility is inferred from a finite Objective value.
# ------------------------------------------------------------------------------
def load_results(path):
    text = path.read_text(
        encoding="utf-8",
        errors="replace",
    )

    cards = re.compile(
        r"<div\s+class=['\"]card['\"]\s+id=['\"](?P<c>c\d+)['\"]>"
        r".*?<pre>(?P<body>.*?)</pre>",
        re.I | re.S,
    )

    obj_re = re.compile(
        r"objective\s*:\s*"
        r"(?P<v>nan|[-+]?\d+(?:\.\d+)?(?:e[-+]?\d+)?)",
        re.I,
    )

    load_re = re.compile(
        r"total_load_mw\s*:\s*"
        r"(?P<v>[-+]?\d[\d,]*(?:\.\d+)?)",
        re.I,
    )

    bus_value_re = re.compile(
        r"^\s*(?P<bus>[A-Za-z0-9_]+)\s*:\s*"
        r"(?P<value>[-+]?\d[\d,]*(?:\.\d+)?)\s*$"
    )

    out = {}

    for match in cards.finditer(text):

        # --------------------------------------------------------------
        # Parse each committed stress result from index.html and store
        # its solver outputs by Commit ID (c1, c2, c3, ...).
        # --------------------------------------------------------------
        commit = match.group("c").lower()
        body = html.unescape(match.group("body"))

        om = obj_re.search(body)
        lm = load_re.search(body)

        if not om:
            continue

        obj = (
            math.nan
            if om.group("v").lower() == "nan"
            else float(om.group("v"))
        )

        bus_import_export = {}
        in_bus_section = False

        for line in body.splitlines():

            if line.startswith(
                "bus_import_export_mw"
            ):
                in_bus_section = True
                continue

            if in_bus_section:

                if "_lmp" in line or line.startswith("lmp_spread"):
                    break

                bm = bus_value_re.match(line)

                if bm:
                    bus_import_export[
                        bm.group("bus")
                    ] = float(
                        bm.group("value").replace(",", "")
                    )

        out[commit] = {
            "objective": obj,
            "load": (
                float(
                    lm.group("v").replace(",", "")
                )
                if lm
                else math.nan
            ),
            "feasible": math.isfinite(obj),
            "bus_import_export": bus_import_export,
        }

    return out

# ------------------------------------------------------------------------------
# save()
# Writes a completed publication artifact as a DPI-controlled PNG into the
# selected output directory and closes the Matplotlib Figure.
# ------------------------------------------------------------------------------
def save(fig, outdir, name):
    outdir.mkdir(parents=True, exist_ok=True)
    path = outdir / name
    fig.savefig(path, dpi=DPI, bbox_inches="tight")
    plt.close(fig)
    print(f"ASR-DBG::Publication figure:\n\t{path}")

# ------------------------------------------------------------------------------
# publication_table()
# Generates an XLSX-defined publication operating-state table for one solved Event.
#
# Reports:
# - Bus generation p_nom
# - Generator marginal cost
# - Bus load p_set
# - Bus net import/export
# - DC_PJM_NE p_nom, marginal cost and p_set
#
# PJM_NE and SERC_SE are highlighted as critical neighboring buses.
#
# main() currently selects the final Event in a Table Artifact's Event grouping
# as the solved operating-state snapshot supplied to this function.
# ------------------------------------------------------------------------------
def publication_table(
    event,
    result,
    outdir,
    artifact,
):

    scenario = artifact["scenario"]
    table_num = artifact["number"]

    fig, ax = plt.subplots(
        figsize=(10.5, 3.4)
    )

    ax.axis("off")

    rows = []

    imports = result.get(
        "bus_import_export",
        {},
    )

    for bus in BUSES:

        asset = event["assets"][bus]

        rows.append(
            [
                bus,
                f"{asset['p_nom']:,.0f}",
                f"{asset['mc']:.0f}",
                f"{asset['load']:,.0f}",
                f"{imports.get(bus, 0.0):+,.0f}",
            ]
        )

    rows.append(
        [
            "DC_PJM_NE",
            f"{event['dc_p_nom']:,.0f}",
            f"{event['dc_mc']:.0f}",
            f"{event['dc_p_set']:,.0f}",
            "—",
        ]
    )

    table = ax.table(
        cellText=rows,
        colLabels=[
            "Bus / Datacenter",
            "p_nom (MW)",
            "MC ($/MWh)",
            "p_set (MW)",
            "Net I/E (MW)\n(+Import / -Export)",
        ],
        loc="center",
        cellLoc="center",
    )

    table.auto_set_font_size(False)
    table.set_fontsize(8.8)
    table.scale(
        1,
        1.35,
    )

    # ------------------------------------------------------------------
    # Highlight critical PJM_NE <-> SERC_SE neighboring buses.
    # ------------------------------------------------------------------
    for table_row, bus in enumerate(
        BUSES,
        start=1,
    ):

        if bus in (
            "PJM_NE",
            "SERC_SE",
        ):

            for col in range(5):

                cell = table[
                    (
                        table_row,
                        col,
                    )
                ]

                cell.set_facecolor(
                    "0.90"
                )

                cell.set_text_props(
                    fontweight="bold"
                )

                cell.set_linewidth(
                    1.5
                )

    ax.set_title(
        f"Table {table_num}. {scenario}",
        fontsize=12,
        pad=12,
    )

    fig.tight_layout()

    # ------------------------------------------------------------------
    # Filename comes directly from XLSX Artifact Name Suffix.
    # ------------------------------------------------------------------
    save(
        fig,
        outdir,
        artifact_filename(
            artifact
        ),
    )

# ------------------------------------------------------------------------------
# render_baseline_chronology()
# Generates the equilibrium-state chronology for an XLSX-defined baseline
# Figure artifact.
#
# Artifact number, Scenario title, filename suffix, Commit source and output
# identity are supplied by devnetDC_sysdsg.xlsx.
# ------------------------------------------------------------------------------
def render_baseline_chronology(
    e,
    events,
    r,
    outdir,
    artifact,
):

    fig = plt.figure(
        figsize=(11.5, 6.2)
    )

    gs = fig.add_gridspec(
        3,
        1,
        height_ratios=[
            1.0,
            0.85,
            0.85,
        ],
        hspace=0.28,
    )

    ax_obj = fig.add_subplot(
        gs[0]
    )

    ax_load = fig.add_subplot(
        gs[1],
        sharex=ax_obj,
    )

    ax_dc = fig.add_subplot(
        gs[2],
        sharex=ax_obj,
    )

    x = [
        0,
        HOURS_PER_YEAR,
    ]

    objective_m = (
        r["objective"]
        / 1e6
    )

    total_load = r["load"]

    # ------------------------------------------------------------------
    # Objective.
    # ------------------------------------------------------------------
    ax_obj.plot(
        x,
        [objective_m] * 2,
        linewidth=2.0,
        label="Objective",
    )

    ax_obj.set_title(
        f"Figure {artifact['number']}. {artifact['scenario']}",
        fontsize=12,
    )

    ax_obj.set_ylabel(
        "Objective\n(USD millions)"
    )

    ax_obj.set_ylim(
        objective_m * 0.97,
        objective_m * 1.05,
    )

    ax_obj.grid(
        True,
        alpha=0.25,
    )

    ax_obj.legend(
        loc="lower center",
        bbox_to_anchor=(
            0.5,
            0.02,
        ),
    )

    # ------------------------------------------------------------------
    # Total system load.
    # ------------------------------------------------------------------
    ax_load.plot(
        x,
        [total_load] * 2,
        linewidth=2.0,
        label="Total System Load",
    )

    ax_load.set_ylabel(
        "System Load\n(MW)"
    )

    ax_load.set_ylim(
        total_load * 0.94,
        total_load * 1.03,
    )

    ax_load.grid(
        True,
        alpha=0.25,
    )

    ax_load.legend(
        loc="lower center",
        bbox_to_anchor=(
            0.5,
            0.02,
        ),
    )

    # ------------------------------------------------------------------
    # Datacenter p_set / BYOG p_nom.
    # ------------------------------------------------------------------
    ax_dc.plot(
        x,
        [e["dc_p_set"]] * 2,
        linewidth=2.0,
        label="DC p_set",
    )

    ax_dc.plot(
        x,
        [e["dc_p_nom"]] * 2,
        linestyle="--",
        linewidth=2.0,
        label="DC BYOG p_nom",
    )

    ax_dc.set_ylabel(
        "Datacenter\n(MW)"
    )

    dc_max = max(
        e["dc_p_set"],
        e["dc_p_nom"],
    )

    ax_dc.set_ylim(
        0,
        dc_max * 1.25,
    )

    ax_dc.set_xlim(
        0,
        HOURS_PER_YEAR,
    )

    ax_dc.set_xlabel(
        "Hour of Year"
    )

    ax_dc.grid(
        True,
        alpha=0.25,
    )

    ax_dc.legend(
        loc="upper right",
        bbox_to_anchor=(
            0.99,
            0.96,
        ),
        ncol=1,
    )

    # ------------------------------------------------------------------
    # Show all active experiment windows against the equilibrium baseline.
    #
    # Event windows provide visual reference for where subsequent Scenario
    # experiments occur in the common 8760-hour chronology. Event conditions
    # are not applied to this baseline Figure.
    # ------------------------------------------------------------------
    for key, event in events.items():

        if key == "baseline":
            continue

        start = event.get(
            "start_hour"
        )

        end = event.get(
            "end_hour"
        )

        if (
            start is None
            or end is None
        ):
            continue

        for ax in (
            ax_obj,
            ax_load,
            ax_dc,
        ):

            ax.axvspan(
                start,
                end,
                alpha=0.12,
                edgecolor="0.35",
                linewidth=0.8,
            )

        ax_obj.text(
            (
                start
                + end
            ) / 2,
            0.96,
            f"E{event['event_num']}",
            transform=ax_obj.get_xaxis_transform(),
            ha="center",
            va="top",
            fontsize=8,
            fontweight="bold",
        )

    plt.setp(
        ax_obj.get_xticklabels(),
        visible=False,
    )

    plt.setp(
        ax_load.get_xticklabels(),
        visible=False,
    )

    fig.tight_layout()

    save(
        fig,
        outdir,
        artifact_filename(
            artifact
        ),
    )

# ------------------------------------------------------------------------------
# format_event_annotation()
# Builds a concise event annotation from narrative XLSX Event Conditions.
#
# Ignores parameter-label rows and retains descriptive condition strings used
# to explain the event beside the chronology plots.
# ------------------------------------------------------------------------------
def format_event_annotation(event):

    lines = [
        f"E{event['event_num']}"
    ]

    for condition in event.get(
        "conditions",
        [],
    ):

        text = str(
            condition.get(
                "name",
                "",
            )
        ).strip()

        if not text:
            continue

        # --------------------------------------------------------------
        # Publication annotations are narrative Event Conditions.
        #
        # Include descriptive conditions containing ":" such as:
        #   Neighbor SERC_SE failure: 20%
        #   DC load reduction: 10%
        #   Transmission Line: L_PJM_NE_SERC_SE failure: 100%
        #   DC BYOG FTM: 25%
        #
        # Ignore parameter-label rows such as:
        #   SERC_SE
        #   L_PJM_NE_SERC_SE
        #   p_set (MW)
        #   byog_p_nom (MW)
        #   Outage (days)
        # --------------------------------------------------------------
        if ":" not in text:
            continue

        lines.append(
            text
        )

    return "\n".join(
        lines
    )

# ------------------------------------------------------------------------------
# extract_percent_event_conditions()
# Extracts percentage-valued narrative Event Conditions from an XLSX Event.
#
# Examples:
# - Neighbor SERC_SE failure:100%                    -> 100
# - Transmission Line: L_PJM_NE_SERC_SE failure:100% -> 100
# - DC BYOG FTM:25%                                  -> 25
#
# Parameter rows without a trailing percentage are ignored.
# Only Event Condition text ending in "<number>%" is interpreted as a driver.
# The extracted numeric value is plotted directly; no semantic transformation
# such as availability = 100 - failure percentage is performed.
# ------------------------------------------------------------------------------
def extract_percent_event_conditions(event):

    conditions = []

    for condition in event.get(
        "conditions",
        [],
    ):

        text = str(
            condition.get(
                "name",
                "",
            )
        ).strip()

        match = re.fullmatch(
            r"(.+):\s*([-+]?\d+(?:\.\d+)?)\s*%",
            text,
        )

        if not match:
            continue

        label = match.group(
            1
        ).strip()

        value = float(
            match.group(
                2
            )
        )

        conditions.append(
            (
                label,
                value,
            )
        )

    return conditions

# ------------------------------------------------------------------------------
# artifact_filename()
# Builds the publication PNG filename from XLSX Artifact metadata.
#
# Uses:
# - Artifact type
# - Artifact number
# - XLSX filename suffix
# Non-alphanumeric characters in the XLSX suffix are normalized to "_".
# The resulting convention is:
#   fig<Number>_<Suffix>.png
#   table<Number>_<Suffix>.png
# ------------------------------------------------------------------------------
def artifact_filename(
    artifact,
):

    prefix = (
        "fig"
        if artifact["type"] == "fig"
        else "table"
    )

    suffix = re.sub(
        r"[^A-Za-z0-9_]+",
        "_",
        artifact["suffix"],
    ).strip("_")

    return (
        f"{prefix}"
        f"{artifact['number']}_"
        f"{suffix}.png"
    )

# ------------------------------------------------------------------------------
# render_event_chronology()
# Generates an XLSX-defined 8760-hour chronology from an Artifact Event grouping.
#
# Builds four aligned panels:
# - Objective / system feasibility
# - Total system load
# - Datacenter p_set / BYOG p_nom
# - Percentage-valued Event Driver conditions
#
# Processing:
# - Replicates Event 0 as the equilibrium baseline across 8760 hours.
# - Overlays each XLSX-defined Event during its Start/End hour window.
# - Maps the Event Commit ID to its solved index.html result.
# - Replaces Objective with NaN during infeasible Events.
# - Marks infeasible intervals with a dark-red "System infeasible" bar.
# - Extracts percentage-valued Event Conditions into the Event Driver panel.
# - Applies the documented Dark Red / Blue Event Driver convention.
#
# Artifact number, title, filename and Event grouping are XLSX-driven.
# ------------------------------------------------------------------------------
def render_event_chronology(
    baseline,
    events,
    results,
    event_keys,
    outdir,
    title,
    filename,
):

    # ------------------------------------------------------------------
    # Publication chronology layout.
    #
    # The bottom panel shows the XLSX Event Conditions that drive each
    # experiment. This keeps experiment input and system response visually
    # separated while sharing the same 8760-hour chronology.
    # ------------------------------------------------------------------
    fig = plt.figure(
        figsize=(11.5, 7.4)
    )

    gs = fig.add_gridspec(
        4,
        1,
        height_ratios=[
            1.0,
            0.85,
            0.85,
            0.70,
        ],
        hspace=0.30,
    )

    ax_obj = fig.add_subplot(
        gs[0]
    )

    ax_load = fig.add_subplot(
        gs[1],
        sharex=ax_obj,
    )

    ax_dc = fig.add_subplot(
        gs[2],
        sharex=ax_obj,
    )

    ax_event = fig.add_subplot(
        gs[3],
        sharex=ax_obj,
    )

    ax_obj.set_title(
        title,
        fontsize=12,
    )

    # ------------------------------------------------------------------
    # Replicate baseline state across 8760 hours.
    # ------------------------------------------------------------------
    baseline_result = results[
        baseline["commit"]
    ]

    objective_base = (
        baseline_result["objective"]
        / 1e6
    )

    load_base = baseline_result[
        "load"
    ]

    hours = list(
        range(
            HOURS_PER_YEAR + 1
        )
    )

    objective = [
        objective_base
    ] * (
        HOURS_PER_YEAR + 1
    )

    system_load = [
        load_base
    ] * (
        HOURS_PER_YEAR + 1
    )

    dc_p_set = [
        baseline["dc_p_set"]
    ] * (
        HOURS_PER_YEAR + 1
    )

    dc_p_nom = [
        baseline["dc_p_nom"]
    ] * (
        HOURS_PER_YEAR + 1
    )

    # ------------------------------------------------------------------
    # XLSX-driven Event Condition chronology.
    #
    # Each distinct percentage-valued Event Condition becomes an
    # independent trace in the bottom event-driver panel.
    # ------------------------------------------------------------------
    event_driver_series = {}

    # ------------------------------------------------------------------
    # Overlay the selected experiment events.
    # ------------------------------------------------------------------
    for key in event_keys:

        event = events[key]

        start = event["start_hour"]
        end = event["end_hour"]

        if (
            start is None
            or end is None
        ):
            continue

        result = results[
            event["commit"]
        ]

        # --------------------------------------------------------------
        # System load and datacenter state exist through the event.
        # --------------------------------------------------------------
        event_load = result.get(
            "load",
            load_base,
        )

        if not math.isfinite(
            event_load
        ):
            event_load = load_base

        for hour in range(
            start,
            end + 1,
        ):

            system_load[hour] = (
                event_load
            )

            dc_p_set[hour] = (
                event["dc_p_set"]
            )

            dc_p_nom[hour] = (
                event["dc_p_nom"]
            )

            if result["feasible"]:

                objective[hour] = (
                    result["objective"]
                    / 1e6
                )

            else:

                objective[hour] = (
                    math.nan
                )

        # --------------------------------------------------------------
        # Populate Event Condition driver traces for this event window.
        # --------------------------------------------------------------
        for (
            condition_label,
            condition_value,
        ) in extract_percent_event_conditions(
            event
        ):

            if condition_label not in event_driver_series:

                event_driver_series[
                    condition_label
                ] = [
                    math.nan
                ] * (
                    HOURS_PER_YEAR + 1
                )

            driver = event_driver_series[
                condition_label
            ]

            for hour in range(
                start,
                end + 1,
            ):
                driver[
                    hour
                ] = condition_value

    # ------------------------------------------------------------------
    # Concise event annotations derived from XLSX Event Conditions.
    # ------------------------------------------------------------------
    annotation_lines = [
        format_event_annotation(
            events[key]
        )
        for key in event_keys
    ]

    annotation_lines = [
        text
        for text in annotation_lines
        if text.strip()
    ]

    if annotation_lines:

        ax_obj.text(
            0.985,
            0.82,
            "\n\n".join(
                annotation_lines
            ),
            transform=ax_obj.transAxes,
            ha="right",
            va="top",
            fontsize=8,
            bbox=dict(
                boxstyle="round,pad=0.45",
                facecolor="white",
                edgecolor="0.65",
                alpha=0.95,
            ),
            zorder=10,
        )

    # ------------------------------------------------------------------
    # Objective chronology.
    # ------------------------------------------------------------------
    ax_obj.plot(
        hours,
        objective,
        linewidth=2.0,
        label="Objective",
    )

    ax_obj.set_ylabel(
        "Objective\n(USD millions)"
    )

    ax_obj.set_ylim(
        objective_base * 0.96,
        objective_base * 1.08,
    )

    ax_obj.grid(
        True,
        alpha=0.25,
    )

    # ------------------------------------------------------------------
    # Total system load chronology.
    # ------------------------------------------------------------------
    ax_load.plot(
        hours,
        system_load,
        linewidth=2.0,
        label="Total System Load",
    )

    ax_load.set_ylabel(
        "System Load\n(MW)"
    )

    load_min = min(
        system_load
    )

    load_max = max(
        system_load
    )

    load_pad = max(
        500.0,
        (
            load_max
            - load_min
        )
        * 0.5,
    )

    ax_load.set_ylim(
        load_min
        - load_pad,
        load_max
        + load_pad,
    )

    ax_load.grid(
        True,
        alpha=0.25,
    )

    ax_load.legend(
        loc="lower center",
        bbox_to_anchor=(
            0.5,
            0.02,
        ),
    )

    # ------------------------------------------------------------------
    # Datacenter p_set / BYOG p_nom chronology.
    # ------------------------------------------------------------------
    ax_dc.plot(
        hours,
        dc_p_set,
        linewidth=2.0,
        label="DC p_set",
    )

    ax_dc.plot(
        hours,
        dc_p_nom,
        linestyle="--",
        linewidth=2.0,
        label="DC BYOG p_nom",
    )

    ax_dc.set_ylabel(
        "Datacenter\n(MW)"
    )

    dc_max = max(
        max(dc_p_set),
        max(dc_p_nom),
    )

    ax_dc.set_ylim(
        0,
        dc_max * 1.25,
    )

    ax_dc.set_xlim(
        0,
        HOURS_PER_YEAR,
    )

    ax_dc.grid(
        True,
        alpha=0.25,
    )

    ax_dc.legend(
        loc="upper right",
        bbox_to_anchor=(
            0.99,
            0.96,
        ),
        ncol=1,
    )

    # ------------------------------------------------------------------
    # XLSX Event Condition driver chronology.
    #
    # Event Driver Line Color Convention
    # ----------------------------------
    # DARK RED = adverse grid/system condition.
    #
    # An Event Condition is classified Dark Red when its label contains:
    #   failure
    #   outage
    #   constraint
    #   deration
    #   congestion
    #
    # BLUE = mitigation/support/default Event Driver.
    #
    # Any percentage-valued Event Condition not matching an adverse keyword
    # is plotted Blue. Current intended Blue drivers include:
    #   DC load reduction
    #   DC BYOG FTM
    #
    # IMPORTANT:
    # Event Condition wording in devnetDC_sysdsg.xlsx is part of this generic
    # classification interface. Keep terminology consistent when adding new
    # Scenarios or Event Driver types.
    # ------------------------------------------------------------------
    for (
        condition_label,
        driver,
    ) in event_driver_series.items():

        label_lower = (
            condition_label.lower()
        )

        is_failure_driver = any(
            keyword in label_lower
            for keyword in (
                "failure",
                "outage",
                "constraint",
                "deration",
                "congestion",
            )
        )

        driver_color = (
            "darkred"
            if is_failure_driver
            else "tab:blue"
        )

        ax_event.plot(
            hours,
            driver,
            linewidth=2.5,
            color=driver_color,
            label=condition_label,
        )

    ax_event.set_ylabel(
        "Event Driver\n(%)"
    )

    ax_event.set_ylim(
        0,
        110,
    )

    ax_event.set_xlim(
        0,
        HOURS_PER_YEAR,
    )

    ax_event.set_xlabel(
        "Hour of Year"
    )

    ax_event.grid(
        True,
        alpha=0.25,
    )

    if event_driver_series:

        ax_event.legend(
            loc="upper right",
            bbox_to_anchor=(
                0.99,
                0.96,
            ),
            ncol=1,
        )

    # ------------------------------------------------------------------
    # Event-window overlays.
    # ------------------------------------------------------------------
    infeasible_legend_added = False
    for key in event_keys:

        event = events[key]

        start = event[
            "start_hour"
        ]

        end = event[
            "end_hour"
        ]

        if (
            start is None
            or end is None
        ):
            continue

        result = results[
            event["commit"]
        ]

        # --------------------------------------------------------------
        # Resolve this event's Commit ID from the XLSX against the
        # corresponding solved result parsed from index.html.
        #
        # Changing the Commit ID in devnetDC_sysdsg.xlsx automatically
        # selects the matching committed solver result for this event.
        # --------------------------------------------------------------
        if result[
            "feasible"
        ]:

            for ax in (
                ax_obj,
                ax_load,
                ax_dc,
                ax_event,
            ):

                ax.axvspan(
                    start,
                    end,
                    alpha=0.10,
                    edgecolor="0.35",
                    linewidth=0.8,
                )

        else:

            # ----------------------------------------------------------
            # Dark-red bar marks an infeasible system interval.
            # ----------------------------------------------------------
            ax_obj.axvspan(
                start,
                end,
                facecolor="darkred",
                alpha=0.75,
                edgecolor="darkred",
                linewidth=1.2,
                label=(
                    "System infeasible"
                    if not infeasible_legend_added
                    else "_nolegend_"
                ),
            )

            infeasible_legend_added = True

        midpoint = (
            start
            + end
        ) / 2

        # --------------------------------------------------------------
        # Event marker.
        # --------------------------------------------------------------
        ax_obj.text(
            midpoint,
            0.96,
            f"E{event['event_num']}",
            transform=ax_obj.get_xaxis_transform(),
            ha="center",
            va="top",
            fontsize=8,
            fontweight="bold",
        )

    # ------------------------------------------------------------------
    # Objective legend includes system infeasibility when present.
    # ------------------------------------------------------------------
    ax_obj.legend(
        loc="lower center",
        bbox_to_anchor=(
            0.5,
            0.02,
        ),
        ncol=2,
    )

    # ------------------------------------------------------------------
    # Hide duplicate X-axis labels.
    # ------------------------------------------------------------------
    plt.setp(
        ax_obj.get_xticklabels(),
        visible=False,
    )

    plt.setp(
        ax_load.get_xticklabels(),
        visible=False,
    )

    plt.setp(
        ax_dc.get_xticklabels(),
        visible=False,
    )

    fig.tight_layout()

    save(
        fig,
        outdir,
        filename,
    )

# ------------------------------------------------------------------------------
# select_publication_artifacts()
# Builds the interactive publication-artifact menu from active XLSX Scenarios.
#
# No Figure/Table numbers or Scenario descriptions are hard-coded here.
# ------------------------------------------------------------------------------
def select_publication_artifacts(
    artifacts,
):

    print(
        "\nSelect publication artifacts to generate:\n"
    )

    print(
        "  0) All publication figures and tables"
    )

    for index, artifact in enumerate(
        artifacts,
        start=1,
    ):

        artifact_type = (
            "Figure"
            if artifact["type"] == "fig"
            else "Table"
        )

        prefix = (
            f"  {index}) "
            f"{artifact_type} "
            f"{artifact['number']} — "
        )

        print(
            textwrap.fill(
                prefix
                + artifact["scenario"],
                width=100,
                subsequent_indent=(
                    " " * len(prefix)
                ),
            )
        )

    print(
        "\nEnter one or more selections separated by commas "
        "[default: 0]:"
    )

    entered = input(
        "> "
    ).strip()

    if (
        not entered
        or entered == "0"
    ):
        return list(
            artifacts
        )

    selected = []

    for value in entered.split(
        ","
    ):

        value = value.strip()

        try:
            index = int(
                value
            )

        except ValueError:
            raise ValueError(
                f"Invalid publication artifact selection: "
                f"{value}"
            )

        if (
            index < 1
            or index > len(
                artifacts
            )
        ):
            raise ValueError(
                f"Publication artifact selection "
                f"out of range: {index}"
            )

        selected.append(
            artifacts[
                index - 1
            ]
        )

    return selected

# ------------------------------------------------------------------------------
# main()
# Orchestrates XLSX-driven publication artifact generation.
#
# Workflow:
# - Load active Events.
# - Load complete publication Artifact definitions.
# - Display the dynamic Artifact-selection menu.
# - Resolve each Artifact's Scenario-specific index.html.
# - Cache parsed stress reports so shared reports are read only once.
# - Validate Artifact Event and Commit references.
# - Route Table Artifacts to publication_table().
# - Route Event-0 Figures to render_baseline_chronology().
# - Route all other Figure Artifacts to render_event_chronology().
#
# No publication Figure/Table numbers or Scenario identities are hard-coded.
# ------------------------------------------------------------------------------
def main():
    args=parse_args()
    try:
        xlsx=Path(args.xlsx).expanduser().resolve()
        print(f"ASR-DBG::Workbook being read:\n\t{xlsx}")
        outdir = Path(
            args.outdir
        ).expanduser().resolve()

        if not xlsx.is_file():
            raise FileNotFoundError(
                f"Workbook not found:\n\t{xlsx}"
            )

        events = load_events(
            xlsx
        )

        artifacts = load_artifacts(
            xlsx
        )

        selected_artifacts = (
            select_publication_artifacts(
                artifacts
            )
        )

        outdir = prepare_output_dir(
            outdir
        )

        # --------------------------------------------------------------
        # Generate XLSX-defined publication artifacts.
        #
        # Each artifact resolves its own devnet_stress.py index.html
        # source from the Scenario definition in devnetDC_sysdsg.xlsx.
        # --------------------------------------------------------------
        results_cache = {}

        for artifact in selected_artifacts:

            if args.html:

                report = Path(
                    args.html
                ).expanduser().resolve()

            else:

                report = resolve_report_path(
                    artifact[
                        "report_path"
                    ],
                    xlsx,
                )

            if not report.is_file():
                raise FileNotFoundError(
                    f"Stress report not found:\n"
                    f"\t{report}\n"
                    f"Scenario:\n"
                    f"\t{artifact['scenario']}"
                )

            report_key = str(
                report
            )

            if report_key not in results_cache:

                results_cache[
                    report_key
                ] = load_results(
                    report
                )

            results = results_cache[
                report_key
            ]

            event_keys = [
                (
                    "baseline"
                    if event_num == 0
                    else f"event{event_num}"
                )
                for event_num in artifact[
                    "events"
                ]
            ]

            missing_events = [
                key
                for key in event_keys
                if key not in events
            ]

            if missing_events:
                raise ValueError(
                    "Artifact references missing events: "
                    + ", ".join(
                        missing_events
                    )
                )

            # ----------------------------------------------------------
            # Validate only commits referenced by this artifact against
            # its Scenario-specific index.html report.
            # ----------------------------------------------------------
            missing_commits = [
                events[key]["commit"]
                for key in event_keys
                if events[key]["commit"] not in results
            ]

            if missing_commits:
                raise ValueError(
                    "Artifact references commits missing from "
                    f"{report.name}: "
                    + ", ".join(
                        missing_commits
                    )
                )

            if artifact[
                "type"
            ] == "table":

                # ------------------------------------------------------
                # For a table spanning multiple events, use the final
                # event as the solved operating-state snapshot.
                # ------------------------------------------------------
                table_event = events[
                    event_keys[-1]
                ]

                publication_table(
                    table_event,
                    results[
                        table_event[
                            "commit"
                        ]
                    ],
                    outdir,
                    artifact,
                )

            elif artifact[
                "type"
            ] == "fig":

                # ------------------------------------------------------
                # Figure rendering is driven by XLSX Artifact metadata.
                #
                # Event 0 alone represents the equilibrium/baseline
                # chronology. All other Figure artifacts are rendered
                # from their explicit XLSX Event grouping.
                # ------------------------------------------------------
                if artifact["events"] == [0]:

                    baseline_event = events[
                        "baseline"
                    ]

                    render_baseline_chronology(
                        baseline_event,
                        events,
                        results[
                            baseline_event[
                                "commit"
                            ]
                        ],
                        outdir,
                        artifact,
                    )

                else:

                    render_event_chronology(
                        baseline=events[
                            "baseline"
                        ],
                        events=events,
                        results=results,
                        event_keys=event_keys,
                        outdir=outdir,
                        title=(
                            f"Figure {artifact['number']}. "
                            f"{artifact['scenario']}"
                        ),
                        filename=artifact_filename(
                            artifact
                        ),
                    )

        print(
            f"ASR-DBG::Selected publication artifacts generated @:\n"
            f"\t{outdir}"
        )

    except Exception as exc:
        print(f"ERROR: {exc}")
        sys.exit(1)

if __name__ == "__main__":
    main()

# ------------------------------------------------------------------------------
# END OF devnet_pub_figs.py
# ------------------------------------------------------------------------------
